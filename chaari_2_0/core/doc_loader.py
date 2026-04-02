# CHAARI 2.0 – core/doc_loader.py — Document Loader & Chunker
# Loads PDF, TXT, MD, PY files and chunks them for RAPTOR tree indexing.

import os
import re
import ast
import logging
from pathlib import Path

from config.rag import (
    CHUNK_SIZE, CHUNK_OVERLAP, CODE_CHUNK_BY_FUNCTION,
    SUPPORTED_EXTENSIONS,
)

logger = logging.getLogger(__name__)



def _extract_pdf(file_path: str) -> str:
    """Extract text from a PDF file."""
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            logger.error("No PDF reader installed. Run: pip install pypdf")
            return ""
    try:
        reader = PdfReader(file_path)
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n\n".join(pages)
    except Exception as e:
        logger.error(f"Failed to read PDF {file_path}: {e}")
        return ""


def _extract_csv(file_path: str) -> str:
    """Extract text from a CSV file — converts rows to readable lines."""
    import csv
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return ""
        header = rows[0] if rows else []
        lines = [", ".join(header)]
        for row in rows[1:]:
            if header:
                parts = [f"{h}: {v}" for h, v in zip(header, row) if v.strip()]
            else:
                parts = [v for v in row if v.strip()]
            if parts:
                lines.append("; ".join(parts))
        return "\n".join(lines)
    except Exception as e:
        logger.error(f"Failed to read CSV {file_path}: {e}")
        return ""


def _extract_docx(file_path: str) -> str:
    """Extract text from a DOCX file."""
    try:
        from docx import Document
    except ImportError:
        logger.error("python-docx not installed. Run: pip install python-docx")
        return ""
    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        logger.error(f"Failed to read DOCX {file_path}: {e}")
        return ""


def _extract_xlsx(file_path: str) -> str:
    """Extract text from an XLSX file — converts sheets/rows to readable text."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return ""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            all_text.append(f"[Sheet: {sheet_name}]")
            header = [str(c) if c is not None else "" for c in rows[0]]
            all_text.append(", ".join(header))
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                if header:
                    parts = [f"{h}: {v}" for h, v in zip(header, cells) if v.strip()]
                else:
                    parts = [v for v in cells if v.strip()]
                if parts:
                    all_text.append("; ".join(parts))
        wb.close()
        return "\n".join(all_text)
    except Exception as e:
        logger.error(f"Failed to read XLSX {file_path}: {e}")
        return ""


def _extract_text(file_path: str) -> str:
    """Extract text from a plain text file."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as e:
        logger.error(f"Failed to read {file_path}: {e}")
        return ""


def _extract_file(file_path: str) -> str:
    """Route file to the correct extractor based on extension."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".csv":
        return _extract_csv(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    elif ext in (".txt", ".md", ".json", ".py"):
        return _extract_text(file_path)
    else:
        logger.warning(f"Unsupported file type: {ext}")
        return ""


def _chunk_python_code(text: str, source: str) -> list[dict]:
    """Chunk Python code by function and class boundaries."""
    chunks = []
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return _chunk_by_tokens(text, source)

    lines = text.split("\n")

    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            start = node.lineno - 1
            end = node.end_lineno if hasattr(node, "end_lineno") and node.end_lineno else start + 20
            chunk_text = "\n".join(lines[start:end])

            if len(chunk_text.split()) < 10:
                continue

            if len(chunk_text.split()) > CHUNK_SIZE:
                sub_chunks = _chunk_by_tokens(chunk_text, source, prefix=f"{node.name}: ")
                chunks.extend(sub_chunks)
            else:
                kind = "class" if isinstance(node, ast.ClassDef) else "function"
                chunks.append({
                    "text": chunk_text,
                    "source": source,
                    "chunk_id": f"{source}::{kind}::{node.name}",
                })

    module_level_lines = []
    top_level_ranges = set()
    for node in ast.iter_child_nodes(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            start = node.lineno - 1
            end = node.end_lineno if hasattr(node, "end_lineno") and node.end_lineno else start + 20
            top_level_ranges.update(range(start, end))

    for i, line in enumerate(lines):
        if i not in top_level_ranges:
            module_level_lines.append(line)

    module_text = "\n".join(module_level_lines).strip()
    if len(module_text.split()) > 15:
        chunks.append({
            "text": module_text,
            "source": source,
            "chunk_id": f"{source}::module_level",
        })

    return chunks if chunks else _chunk_by_tokens(text, source)


def _chunk_by_tokens(text: str, source: str, prefix: str = "") -> list[dict]:
    """Split text into chunks of ~CHUNK_SIZE words with overlap."""
    words = text.split()
    if not words:
        return []

    chunks = []
    idx = 0
    chunk_num = 0

    while idx < len(words):
        end = min(idx + CHUNK_SIZE, len(words))
        chunk_text = " ".join(words[idx:end])

        if len(chunk_text.split()) >= 10:  
            chunks.append({
                "text": f"{prefix}{chunk_text}" if prefix else chunk_text,
                "source": source,
                "chunk_id": f"{source}::chunk_{chunk_num}",
            })
            chunk_num += 1

        idx += CHUNK_SIZE - CHUNK_OVERLAP
        if idx >= len(words):
            break

    return chunks


def load_and_chunk_file(file_path: str) -> list[dict]:
    """
    Load a single file and return chunks.

    Returns:
        List of dicts: [{"text": str, "source": str, "chunk_id": str}, ...]
    """
    ext = Path(file_path).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        logger.warning(f"Skipping unsupported file: {file_path}")
        return []

    text = _extract_file(file_path)
    if not text or len(text.strip()) < 20:
        return []

    source = os.path.basename(file_path)

    if ext == ".py" and CODE_CHUNK_BY_FUNCTION:
        return _chunk_python_code(text, source)

    return _chunk_by_tokens(text, source)


def load_and_chunk_directory(dir_path: str, recursive: bool = True) -> list[dict]:
    """
    Load all supported files from a directory and return chunks.

    Returns:
        List of all chunks from all files.
    """
    all_chunks = []
    dir_path = Path(dir_path)

    if not dir_path.exists():
        logger.error(f"Directory not found: {dir_path}")
        return []

    pattern = "**/*" if recursive else "*"
    for file_path in sorted(dir_path.glob(pattern)):
        if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_EXTENSIONS:
            if "__pycache__" in str(file_path) or file_path.name.startswith("."):
                continue
            chunks = load_and_chunk_file(str(file_path))
            all_chunks.extend(chunks)
            if chunks:
                logger.info(f"  Loaded {file_path.name}: {len(chunks)} chunks")

    return all_chunks
